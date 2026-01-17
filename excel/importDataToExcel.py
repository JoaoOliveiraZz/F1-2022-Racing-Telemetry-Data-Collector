import os
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL_PATH = os.path.join(BASE_DIR, "telemetryData.xlsx")

def append_to_excel(df, filename=DEFAULT_EXCEL_PATH, sheet_name="sheet1"):
    if not os.path.exists(filename):
        df.to_excel(filename, sheet_name=sheet_name, index=False)
        return

    # Descobre a última linha
    book = load_workbook(filename)

    if sheet_name in book.sheetnames:
        start_row = book[sheet_name].max_row
    else:
        start_row = 0

    with pd.ExcelWriter(
        filename,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
    ) as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=start_row,
            index=False,
            header=False
        )
